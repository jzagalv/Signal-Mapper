from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from domain.services.pending_service import count_pending_for_bay, count_pending_for_device
from domain.services.interlock_service import interlock_tags


HEADERS = [
    "Equipo",
    "Tipo Equipo",
    "Dir",
    "Señal ID",
    "Nombre Señal",
    "Texto (Desde/Hacia)",
    "Naturaleza",
    "Estado",
    "Block Pruebas (OUT)",
    "Enclavamientos (IN)",
]


def export_project_to_excel(project, path: str) -> None:
    """Exporta un Excel con:
    - 1 hoja 'Resumen' (conteos por bahía/equipo)
    - 1 hoja por bahía (detalles de señales)

    Nota de ingeniería:
    - Block de pruebas sólo aplica a OUT.
    - Enclavamientos sólo aplican a IN.
    """
    wb = Workbook()

    # Resumen
    ws_sum = wb.active
    ws_sum.title = "Resumen"
    _build_summary_sheet(ws_sum, project)

    # Hojas por bahía
    for bay_id, bay in project.bays.items():
        ws = wb.create_sheet(title=_safe_sheet_name(bay.name or bay_id))

        ws.append(["Bahía", bay.name or bay_id])
        ws.append([])

        ws.append(HEADERS)
        header_row = ws.max_row
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for dev in bay.devices.values():
            ws.append([])
            ws.append([dev.name, dev.dev_type, "", "", "", "", "", "", "", ""])
            dev_row = ws.max_row
            for c in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=dev_row, column=c)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left" if c == 1 else "center")

            ends = [("IN", e) for e in dev.inputs] + [("OUT", e) for e in dev.outputs]
            for direction, e in ends:
                sig = bay.signals.get(e.signal_id)
                sig_name = sig.name if sig else e.signal_id
                nature = sig.nature if sig else "DIGITAL"

                # Block de pruebas sólo OUT
                test_block = "Sí" if (direction == "OUT" and bool(getattr(e, "test_block", False))) else ""

                # Enclavamientos sólo IN
                interlocks = "; ".join(interlock_tags(getattr(e, "interlocks", None))) if direction == "IN" else ""

                ws.append(
                    [
                        "",
                        "",
                        direction,
                        e.signal_id,
                        sig_name,
                        e.text or "",
                        nature,
                        e.status,
                        test_block,
                        interlocks,
                    ]
                )
                row = ws.max_row
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=9).alignment = Alignment(horizontal="center")

        _autosize(ws)
        ws.freeze_panes = "A4"

    wb.save(path)


def _build_summary_sheet(ws, project) -> None:
    ws.append(["Proyecto", getattr(project, "name", "")])
    ws.append([])

    ws.append(["Resumen por bahía"])
    ws.append(["Bahía", "Pendientes Total", "Pendientes OUT", "Pendientes IN", "Equipos"])
    for c in range(1, 6):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

    for bay_id, bay in project.bays.items():
        counts = count_pending_for_bay(bay)
        ws.append([bay.name or bay_id, counts["total_pending"], counts["out_pending"], counts["in_pending"], len(bay.devices)])

    ws.append([])
    ws.append(["Detalle por equipo"])
    ws.append(["Bahía", "Equipo", "Tipo", "Pendientes Total", "Pendientes OUT", "Pendientes IN", "Total IN", "Total OUT"])
    for c in range(1, 9):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)

    for bay_id, bay in project.bays.items():
        for dev in bay.devices.values():
            pc = count_pending_for_device(dev)
            ws.append(
                [
                    bay.name or bay_id,
                    dev.name,
                    dev.dev_type,
                    pc["total_pending"],
                    pc["out_pending"],
                    pc["in_pending"],
                    len(dev.inputs),
                    len(dev.outputs),
                ]
            )

    _autosize(ws)


def _safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    s = "".join(ch for ch in name if ch not in bad)
    return (s or "Hoja")[:31]


def _autosize(ws) -> None:
    # auto ancho básico
    from openpyxl.utils import get_column_letter

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 80)
